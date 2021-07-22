"""
Perform politeness research in tweets
THis version created by Erwin R. Komen
Date: 14/jul/2021
"""

import sys, getopt, os.path, importlib
import os, sys
import csv, json
import openpyxl
from openpyxl import Workbook

# My own stuff
import utils
from TextSearch import TextSearch
# The data for this project
import data
from pattern.text.nl import parse, split

# Generic error handling
errHandle = utils.ErrHandle()


# ----------------------------------------------------------------------------------
# Name :    main
# Goal :    Main body of the function
# History:
# 15/sep/2017    ERK Created
# ----------------------------------------------------------------------------------
def main(prgName, argv) :
    flInput = ''        # input directory name
    flOutput = ''       # output directory name
    method = "beleefd"

    try:
        sSyntax = prgName + '[-m <method>] -i <input file>'
        # get all the arguments
        try:
            # Get arguments and options
            opts, args = getopt.getopt(argv, "hm:i:", ["-method=", "-inputfile="])
        except getopt.GetoptError:
            print(sSyntax)
            sys.exit(2)
        # Walk all the arguments
        for opt, arg in opts:
            if opt in ("-h", "--help"):
                print(sSyntax)
                sys.exit(0)
            elif opt in ("-i", "--ifile"):
                flInput = arg
            elif opt in ("-m", "--method"):
                method = arg
        # Check if all arguments are there
        if (flInput == ''):
            errHandle.DoError(sSyntax)

        # Determine what the output file will be
        flOutput = flInput.replace(".xlsx", "_auto.xlsx")

        # Continue with the program
        errHandle.Status('Input is "' + flInput + '"')
        errHandle.Status('Output is "' + flOutput + '"')
        errHandle.Status('Method is "' + method + '"')

        # Call the function that does the job
        oArgs = {'input': flInput, 'output': flOutput}

        # Action depends on method
        if method == "beleefd":
            # Politeness research
            if (not process_beleefd(oArgs)) :
                errHandle.DoError("Could not complete")
                return False
        elif method == "frequency":
            # Frequency analysis
            if (not process_frequency(oArgs)):
                errHandle.DoError("Could not complete")
                return False

    
        # All went fine  
        errHandle.Status("Ready")
    except:
        # act
        errHandle.DoError("main")
        return False

# ----------------------------------------------------------------------------------
# Name :    process_beleefd
# Goal :    Politeness research
# History:
# 14/jul/2021    ERK Created
# ----------------------------------------------------------------------------------
def process_beleefd(oArgs):
    """Process one lingopart file"""

    # Initialisations
    method = "excel_to_excel"
    method = "excel_to_json"
    lOutput = []
    tweet_cols = {
        "Text_tweet_1": {"src": 0, "dst": 0}, 
        "Text_tweet_2": {"src": 0, "dst": 0}, 
        "Text_tweet_3": {"src": 0, "dst": 0}, 
        "Text_tweet_4": {"src": 0, "dst": 0}, 
        "Text_tweet_5": {"src": 0, "dst": 0},
        "Text_tweet_6": {"src": 0, "dst": 0}, 
        "Text_tweet_7": {"src": 0, "dst": 0}, 
        "Text_tweet_8": {"src": 0, "dst": 0}, 
        "Text_tweet_9": {"src": 0, "dst": 0}
        }

    try:
        # Recover the arguments
        if "input" in oArgs: flInput = oArgs["input"]
        if "output" in oArgs: flOutput = oArgs["output"]

        # Check input file
        if not os.path.isfile(flInput):
            errHandle.Status("Please specify an input FILE")
            return False

        # Open the Excel file
        wb = openpyxl.load_workbook(flInput, read_only=True)
        ws = wb.active

        if method == "excel_to_excel":
            # Open an output Excel file
            wb_dst = openpyxl.Workbook()
            ws_dst = wb_dst.active

        # Initialize the evaluation functions
        for oRule in data.auto_rules:
            oRule['textsearch'] = None
            oRule['textsearch'] = TextSearch(oRule['words'])

        if method == "excel_to_json":
            # Create a row for the header
            lRow = []
        # The first 11 headers need to be copied
        row_no = 1
        col_no = 1
        col_no_dst = 1
        max_col = col_no
        while col_no < 130 or ws.cell(row=row_no, column=col_no).value != None:
            cell = ws.cell(row=row_no, column=col_no)
            if cell.value != None:
                # Get the name of this column
                col_name = cell.value
                
                # Always copy the header
                if method == "excel_to_json":
                    lRow.append(col_name)
                else:
                    ws_dst.cell(row=row_no, column=col_no_dst).value = col_name
                if col_name in tweet_cols.keys():
                    # Get this object
                    oTweetCol = tweet_cols[col_name]
                    # Set the source number of this column
                    oTweetCol['src'] = col_no
                    oTweetCol['dst'] = col_no_dst
                    # Insert columns as per the data (data labels)
                    for oLabel in data.auto_rules:
                        label = oLabel['label']
                        col_no_dst += 1
                        # Name the column                        
                        if method == "excel_to_json":
                            lRow.append(label)
                        else:
                            ws_dst.cell(row=row_no, column=col_no_dst).value = label

            # Advance the column numbers
            max_col = col_no
            col_no += 1
            col_no_dst += 1
        # Check the first row
        if method == "excel_to_json":
            lOutput.append(lRow)
        else:
            first_row = ws_dst[1]
            headers = [x.value for x in first_row]

        # Walk the remaining rows
        tweet_cols_src = [x['src'] for key, x in tweet_cols.items()]
        row_no += 1
        for row in ws.iter_rows(min_row=row_no, max_col=max_col):
            # while ws.cell(row=row_no, column=1).value != None:
            # Show where we are
            # errHandle.Status("Processing row: {}, iD={}".format(row_no, ws.cell(row=row_no, column=1).value))
            errHandle.Status("Processing row: {}, iD={}".format(row_no, row[0].value))
            if method == "excel_to_json":
                # Create an object row for this excel row
                lRow = []
            # Walk the columns, copying as neeeded
            col_no = 1
            col_no_dst = 1
            # while col_no < 130 or ws.cell(row=row_no, column=col_no).value != None:
            for cell in row:
                # Always copy the cell contents
                # cell_value = ws.cell(row=row_no, column=col_no).value
                cell_value = cell.value
                if method == "excel_to_json":
                    if cell_value == None:
                        lRow.append(None)
                    else:
                        lRow.append("{}".format(cell_value))
                else:
                    ws_dst.cell(row=row_no, column=col_no_dst).value = cell_value

                # Do we need to do some analysis here?
                if col_no in tweet_cols_src:
                    # Yes! get the text of this cell
                    sText = "{}".format(cell_value)
                    # Walk all the auto_rules
                    for oLabel in data.auto_rules:
                        # Advance the destination column
                        col_no_dst += 1
                        # Calculate the value for this evaluation
                        evaluation = 1 if oLabel['textsearch'].exists(sText) else 0
                        if method == "excel_to_json":
                            lRow.append(evaluation)
                        else:
                            ws_dst.cell(row=row_no, column=col_no_dst).value = evaluation

                # Advance the column numbers
                col_no += 1
                col_no_dst += 1

            if method == "excel_to_json":
                lOutput.append(lRow)
            # Go to the next row
            row_no += 1

        # Save the output
        if method == "excel_to_json":
            # Save the JSON
            errHandle.Status("Writing json...")
            flJson = flOutput.replace(".xlsx", ".json")
            with open(flJson, "w") as fp:
                json.dump(lOutput, fp, indent=2)
            # Copy everything into a new Excel
            errHandle.Status("Converting Json to Excel...")
            wb_dst = openpyxl.Workbook()
            ws_dst = wb_dst.active
            row_no = 1
            for lRow in lOutput:
                for idx, value in enumerate(lRow):
                    col_no = idx + 1
                    if value != None:
                        ws_dst.cell(row=row_no, column=col_no).value = value
                # Go to the next row
                row_no += 1

            # Save this excel
            errHandle.Status("Writing Excel...")
            wb_dst.save(flOutput)
        else:
            wb_dst.save(flOutput)

        # Return correctly
        return True
    except:
        msg = errHandle.get_error_message()
        errHandle.DoError("process_beleefd")
        return False

# ----------------------------------------------------------------------------------
# Name :    process_frequency
# Goal :    Frequency research
# History:
# 19/jul/2021    ERK Created
# ----------------------------------------------------------------------------------
def process_frequency(oArgs):
    """Process one lingopart file"""

    method = "none"
    row_nld_first = 64
    col_klacht = 3
    col_service = 5
    col_language = 22
    max_col = 22
    min_row = 3
    max_row = 300
    lst_research = [
        {'label': 'uitleg',                         'col': 16, 'text': ""},
        {'label': 'reverse blame',                  'col': 17, 'text': ""},
        {'label': 'herhaling positieve elementen',  'col': 19, 'text': ""},
        {'label': 'downtoanen',                     'col': 20, 'text': ""},
        ]
    skippable_pos = [".", ",", "DT"]
    bLemma = False

    try:
        # Recover the arguments
        if "input" in oArgs: flInput = oArgs["input"]
        if "output" in oArgs: flOutput = oArgs["output"]

        # Check input file
        if not os.path.isfile(flInput):
            errHandle.Status("Please specify an input FILE")
            return False

        # Open the Excel file
        wb = openpyxl.load_workbook(flInput, read_only=True)
        ws = wb.active

        row_no = min_row
        for row in ws.iter_rows(min_row=min_row, max_row=max_row, max_col=max_col):
            # Show where we are
            errHandle.Status("Processing row {}".format(row_no))
            # Get the language
            language = "{}".format(row[col_language-1].value)
            if language == "nl":
                # Get the text
                sText = "{}".format(row[col_service-1].value)

                # Visit each of the lists
                for oResearch in lst_research:
                    col_idx = oResearch['col'] - 1
                    # Get the value
                    cell = row[col_idx]
                    cell_value = "{}".format(cell.value)
                    if cell_value == "1":
                        # Add this one to the list
                        oResearch['text'] += sText + " "

            row_no += 1
        # Close the workbook
        wb.close()    

        flOutput = flOutput.replace("_auto.xlsx", "_freq.xlsx")
        wb = Workbook()
        ws = wb.active

        # Now process the research areas
        col_no = 1
        for oResearch in lst_research:
            # Get the text
            sText = oResearch['text']
            # Tokenize and get lemmata
            arTok = parse(sText, tokenize=True, lemmata=True)
            # Create a frequency list
            oFreq = {}
            for snt in split(arTok):
                for oWord in snt.words:
                    if not oWord.pos in skippable_pos:
                        if bLemma:
                            sWord = "{}_{}".format(oWord.lemma, oWord.pos)
                        else:
                            sWord = "{}_{}".format(oWord.string, oWord.pos)
                        if not sWord in oFreq:
                            oFreq[sWord] = 1
                        else:
                            oFreq[sWord] += 1
            lstFreq = [{'freq': v, 'key': k} for k,v in oFreq.items()]
            lstFreq = sorted(lstFreq, key=lambda x: x['freq'], reverse=True)
            oResearch['freq'] = lstFreq
            # Set the column header
            row_no = 1
            ws.cell(row_no, col_no).value = oResearch['label']
            ws.cell(row_no, col_no+1).value = "frequency"
            for idx, oFreq in enumerate(lstFreq):
                row_no += 1
                ws.cell(row_no, col_no).value = oFreq['key']
                ws.cell(row_no, col_no+1).value = oFreq['freq']

            # Go to the next column
            col_no += 2

        wb.save(flOutput)

        return True
    except:
        msg = errHandle.get_error_message()
        errHandle.DoError("process_beleefd")
        return False


# ----------------------------------------------------------------------------------
# Goal :  If user calls this as main, then follow up on it
# ----------------------------------------------------------------------------------
if __name__ == "__main__":
  # Call the main function with two arguments: program name + remainder
  main(sys.argv[0], sys.argv[1:])
